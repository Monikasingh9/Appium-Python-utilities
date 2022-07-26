import logging
import time
import re
from appium import webdriver
from configparser import ConfigParser
import openpyxl
from appium.webdriver.common.appiumby import AppiumBy
from Var_message import message_app, message_activity, message_status, message_sender_name, message_recipient_number, \
    send_message, input_message, new_message_app, compose_message_text, message_text_locator, add_ppl, more_options, \
    next_locator, skip_locator

from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import ElementNotVisibleException, NoSuchElementException,ElementClickInterceptedException

"""
device can take the values based on config.ini e.g.,device1,device3 etc
similarly sheetname can take the values based on config.ini eg.,device1,device3 etc
"""

# class message_app():
config = ConfigParser()
config.read("./Resources/config.ini")
myconfig = {}
myconfig['server'] = config.get('Server', 'server_url')
myconfig['device1'] = config.get('Device', 'device1')
myconfig['device2'] = config.get('Device', 'device2')
myconfig['device3'] = config.get('Device', 'device3')
myconfig['excel_file'] = config.get('DataFile', 'excel')
myconfig['phone_number1'] = config.get('Phone Numbers', 'device1')
myconfig['phone_number2'] = config.get('Phone Numbers', 'device2')
myconfig['phone_number3'] = config.get('Phone Numbers', 'device3')


# def __init__(self):
#     pass

def explicitly_wait(driver,time_out,element_loc):
    wait=WebDriverWait(driver,time_out,poll_frequency=1,ignored_exceptions=[NoSuchElementException,ElementNotVisibleException])
    my_element=wait.until(lambda x:x.find_element(AppiumBy.ID,element_loc))
    return my_element

def get_desired_capabilities(device):
    workbook = openpyxl.load_workbook(myconfig['excel_file'])
    sheet_name = myconfig[device]
    sheet = workbook[sheet_name]
    totalrows = sheet.max_row
    totalcolumns = sheet.max_column
    capabilities = {}
    for c in range(1, totalcolumns + 1):
        mykey = sheet.cell(1, c).value
        myvalue = sheet.cell(2, c).value
        capabilities[mykey] = myvalue
    return capabilities


def get_driver(device):
    desired_caps = get_desired_capabilities(device)
    server = myconfig['server']
    driver = webdriver.Remote(server, desired_caps)
    driver.implicitly_wait(5)
    return driver


def open_message_app(device):
    driver = get_driver(device)
    dl = driver.start_activity(message_app,message_activity)
    dl.implicitly_wait(10)
    return dl


def send_group_message(device_from,device1,device2):
    driver = open_message_app('device' + str(device_from))
    try:
        ele = explicitly_wait(driver, 15,new_message_app)
        ele.click()
    except:
        logging.error('element not found')
    # driver.find_element(by=AppiumBy.ID, value=new_message_app).click()
    dl = driver.find_element(by=AppiumBy.ID, value=message_recipient_number)
    if dl.is_displayed():
        logging.info('New Massage page is open')
    else:
        logging.error('New message page could not be opened')

    ph_num1=myconfig['phone_number' + str(device1)]
    ph_num2=myconfig['phone_number'+str(device2)]
    dl.send_keys(ph_num1)
    driver.execute_script('mobile: performEditorAction', {'action': 'done'})
    driver.find_element(AppiumBy.ID,value=more_options).click()
    driver.find_element(AppiumBy.XPATH,value="//android.widget.TextView[@text='Details']").click()
    driver.find_element(AppiumBy.ID,value=add_ppl).click()
    driver.find_element(by=AppiumBy.ID, value=message_recipient_number).send_keys(ph_num2)
    driver.execute_script('mobile: performEditorAction', {'action': 'done'})
    driver.find_element(AppiumBy.XPATH,next_locator).click()
    driver.find_element(AppiumBy.XPATH,skip_locator).click()
    # ele1 = explicitly_wait(driver, 5,'XPATH', next_locator)
    # ele1.click()
    # ele2 = explicitly_wait(driver, 5,'XPATH', skip_locator)
    # ele2.click()
    ele3 = explicitly_wait(driver, 5,compose_message_text)
    ele3.send_keys(input_message)
    driver.find_element(by=AppiumBy.ID, value=send_message).click()
    time.sleep(5)
    d = driver.find_element(by=AppiumBy.ID, value=message_status).text

    if d.__contains__('Delivered') or d.__contains__('Sent') or d.__contains__('Now • MMS'):
        logging.info('The message is delivered')
    else:
        logging.error('Message not sent')
    return driver



def send_a_message(device_from, device_to):
    driver = open_message_app('device' + str(device_from))
    try:
        ele=explicitly_wait(driver,15,new_message_app)
        ele.click()
    except:
        pass
    
    dl = driver.find_element(by=AppiumBy.ID, value=message_recipient_number)
    if dl.is_displayed():
        logging.info('New Massage page is open')
    else:
        logging.error('New message page could not be opened')
    ph_num = myconfig['phone_number' + str(device_to)]
    dl.send_keys(ph_num)
    driver.execute_script('mobile: performEditorAction', {'action': 'done'})
    dl = driver.find_element(by=AppiumBy.ID, value=compose_message_text)
    dl.send_keys(input_message)
    driver.find_element(by=AppiumBy.ID,value=send_message).click()
    time.sleep(5)
    d = driver.find_element(by=AppiumBy.ID, value=message_status).text
    if d.__contains__('Delivered') or d.__contains__('Sent') or d.__contains__('sending'):
        logging.info('The message is delivered')
    else:
        logging.error('Message not sent')
    return driver


def check_message(device_from, device_to):
    driver = open_message_app('device' + str(device_to))
    ph_num = myconfig['phone_number' + str(device_from)]
    msg_list = driver.find_elements(by=AppiumBy.ID, value=message_sender_name)
    cnt = len(msg_list)
    for i in range(cnt):
        s = msg_list[i].text
        s1 = re.sub('[()-]', "", s)
        s1 = s1.replace(" ", "")
        if s1 == ph_num:
            msg_list[i].click()
            new_list = driver.find_elements(by=AppiumBy.ID,
                                            value=message_text_locator)
            ct = len(new_list)

            if new_list[ct - 1].text == input_message:
                logging.info('Message Received successfully')
            else:
                logging.error('Message was not received')

            break
    return driver

def test_teardown(driver1,driver2):
    driver1.terminate_app(message_app)
    driver2.terminate_app(message_app)
    driver1.quit()
    driver2.quit()


