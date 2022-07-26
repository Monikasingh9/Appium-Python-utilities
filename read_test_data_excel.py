import openpyxl

# get the list of phone numbers to be tested from excel file

def get_phone_numbers_from_excel(excel_file, sheet_name):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheet_name]
    totalrows = sheet.max_row
    totalcoloumns = sheet.max_column
    number_list = []
    print(f'total no of rows is {totalrows} and total no of columns is {totalcoloumns}')
    for r in range(2, totalrows + 1):
        number_list.append(sheet['A' + str(r)].value)
    return number_list

# get the desired capabilities from excel file

def get_capabilities(excel_file, sheet_name, dev_no):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheet_name]
    totalrows = sheet.max_row
    totalcolumns = sheet.max_column
    cap = []
    capabilities = ''
    use_row = 1
    for c in range(1, totalcolumns + 1):
        title = sheet.cell(1, c).value
        print(title)
        if title == 'deviceNo':
            for r in range(2, totalrows, 1):
                num = sheet.cell(r, c).value
                print(num)
                if num == dev_no:
                    use_row = r

    for c in range(1, totalcolumns + 1):
        title = sheet.cell(1, c).value
        content = sheet.cell(use_row, c).value
        if title != 'deviceNo':
            cap.append(str(title) + "=\'" + str(content) + "\'")

    capabilities = ','.join(cap)
    return capabilities

